"""
Document loading and chunking utilities with OCR support.

Supports:
  Text-based:  PDF, TXT, DOCX, CSV, Markdown, XLSX
  Image/OCR:   PNG, JPG, JPEG, TIFF, BMP, GIF, WEBP
  Scanned PDFs are auto-detected and OCR'd.
"""

import os
import tempfile
from pathlib import Path
from typing import List

from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
    Docx2txtLoader,
    CSVLoader,
    UnstructuredMarkdownLoader,
)

from config import (
    CHUNK_SIZE,
    CHUNK_OVERLAP,
    SUPPORTED_EXTENSIONS,
    TEXT_EXTENSIONS,
    IMAGE_EXTENSIONS,
    TESSERACT_LANG,
    OCR_DPI,
)

# ── Lazy imports for OCR (only loaded when needed) ───────────

_pytesseract = None
_Image = None
_convert_from_path = None


def _ensure_ocr_imports():
    """Lazily import OCR dependencies so the app starts fast
    even if Tesseract is not installed."""
    global _pytesseract, _Image, _convert_from_path
    if _pytesseract is None:
        try:
            import pytesseract
            from PIL import Image
            from pdf2image import convert_from_path

            _pytesseract = pytesseract
            _Image = Image
            _convert_from_path = convert_from_path
        except ImportError as e:
            raise ImportError(
                "OCR dependencies missing. Install them:\n"
                "  pip install pytesseract Pillow pdf2image\n"
                "  Also install Tesseract: brew install tesseract (macOS)"
            ) from e


# ── Text-based loaders ────────────────────────────────────────

LOADER_MAP = {
    ".pdf": PyPDFLoader,
    ".txt": TextLoader,
    ".docx": Docx2txtLoader,
    ".csv": CSVLoader,
    ".md": UnstructuredMarkdownLoader,
}


def _load_xlsx(file_path: str) -> List[Document]:
    """Load an Excel file and return Documents (one per sheet)."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    docs: List[Document] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append("\t".join(str(c) if c is not None else "" for c in row))
        text = "\n".join(rows)
        if text.strip():
            docs.append(
                Document(
                    page_content=text,
                    metadata={"source": Path(file_path).name, "sheet": sheet_name},
                )
            )
    return docs


# ── OCR loaders ───────────────────────────────────────────────


def _ocr_image(file_path: str) -> List[Document]:
    """Run Tesseract OCR on an image file."""
    _ensure_ocr_imports()
    img = _Image.open(file_path)
    text = _pytesseract.image_to_string(img, lang=TESSERACT_LANG)
    if not text.strip():
        return []
    return [
        Document(
            page_content=text,
            metadata={"source": Path(file_path).name, "loader": "ocr_image"},
        )
    ]


def _ocr_pdf(file_path: str) -> List[Document]:
    """Convert each PDF page to image, then OCR."""
    _ensure_ocr_imports()
    images = _convert_from_path(file_path, dpi=OCR_DPI)
    docs: List[Document] = []
    for i, img in enumerate(images):
        text = _pytesseract.image_to_string(img, lang=TESSERACT_LANG)
        if text.strip():
            docs.append(
                Document(
                    page_content=text,
                    metadata={
                        "source": Path(file_path).name,
                        "page": i + 1,
                        "loader": "ocr_pdf",
                    },
                )
            )
    return docs


def _is_scanned_pdf(docs: List[Document]) -> bool:
    """Heuristic: if PyPDF extracted very little text, it's likely scanned."""
    total_text = sum(len(d.page_content.strip()) for d in docs)
    return total_text < 50  # Less than 50 chars across all pages → scanned


# ── Main loading functions ────────────────────────────────────


def load_single_file(file_path: str) -> List[Document]:
    """Load a single file. Auto-selects OCR for images and scanned PDFs."""
    ext = Path(file_path).suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type: {ext}. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    # ── Image files → always OCR ──────────────────────────────
    if ext in IMAGE_EXTENSIONS:
        return _ocr_image(file_path)

    # ── Excel ─────────────────────────────────────────────────
    if ext == ".xlsx":
        return _load_xlsx(file_path)

    # ── PDF: try text extraction first, fall back to OCR ──────
    if ext == ".pdf":
        loader = PyPDFLoader(file_path)
        docs = loader.load()
        if _is_scanned_pdf(docs):
            # Scanned/handwritten PDF → OCR
            docs = _ocr_pdf(file_path)
        for doc in docs:
            doc.metadata["source"] = Path(file_path).name
        return docs

    # ── Other text-based files ────────────────────────────────
    loader = LOADER_MAP[ext](file_path)
    docs = loader.load()
    for doc in docs:
        doc.metadata["source"] = Path(file_path).name
    return docs


def load_files(file_paths: List[str]) -> List[Document]:
    """Load multiple files and return a combined list of Documents."""
    all_docs: List[Document] = []
    errors: List[str] = []
    for fp in file_paths:
        try:
            all_docs.extend(load_single_file(fp))
        except Exception as e:
            errors.append(f"{Path(fp).name}: {e}")
    return all_docs, errors


def chunk_documents(documents: List[Document]) -> List[Document]:
    """Split documents into overlapping chunks optimized for executive analysis."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", "; ", ", ", " ", ""],
    )
    return splitter.split_documents(documents)
